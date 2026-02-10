"""Import historical data from the scraped Excel spreadsheet into Elasticsearch.

Usage (run inside the backend container or locally with ES running):
    python import_spreadsheet.py /path/to/polymarket_snapshots.xlsx

Imports from these sheets:
  - snapshots_wide: historical price snapshots (wide format)
  - markets: market metadata
  - tracked_markets: tracking configuration
"""

import asyncio
import json
import logging
import sys
from datetime import datetime, timezone

import openpyxl
from elasticsearch import AsyncElasticsearch
from elasticsearch.helpers import async_bulk

import config
from utils.dedup import generate_snapshot_doc_id

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

ES_HOST = config.ES_HOST


def parse_number(val) -> float:
    """Parse a number that may use comma as decimal separator."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    # Handle European comma decimal: "68338,18785" → 68338.18785
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    return float(s)


def read_snapshots_wide(wb) -> list[dict]:
    """Read the snapshots_wide sheet into snapshot documents."""
    ws = wb["snapshots_wide"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    logger.info("snapshots_wide headers: %s", headers)
    docs = []

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        ts_raw = str(row[0])
        market_id = str(int(row[1])) if isinstance(row[1], float) else str(row[1])
        question = str(row[2] or "")
        yes_price = parse_number(row[3])
        no_price = parse_number(row[4])
        # yes_cents/no_cents in sheet are like "2¢", "98¢" — recompute from prices
        yes_cents = round(yes_price * 100)
        no_cents = round(no_price * 100)
        spread_raw = parse_number(row[7])
        # Some sheets have spread=1.0 as a placeholder; recompute
        spread = round(abs(yes_price - no_price), 6)
        volume = parse_number(row[8])
        liquidity = parse_number(row[9])
        active = row[10] if isinstance(row[10], bool) else str(row[10]).lower() == "true" if row[10] else True
        closed = row[11] if isinstance(row[11], bool) else str(row[11]).lower() == "true" if row[11] else False
        market_slug = str(row[12] or "") if len(row) > 12 else ""

        # Parse timestamp
        try:
            if isinstance(ts_raw, datetime):
                ts = ts_raw.replace(tzinfo=timezone.utc) if ts_raw.tzinfo is None else ts_raw
            else:
                ts_str = ts_raw.replace("+00:00", "Z").rstrip("Z")
                # Try parsing with microseconds
                for fmt in ["%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"]:
                    try:
                        ts = datetime.strptime(ts_str, fmt).replace(tzinfo=timezone.utc)
                        break
                    except ValueError:
                        continue
                else:
                    logger.warning("Cannot parse timestamp: %s", ts_raw)
                    continue
        except Exception as e:
            logger.warning("Timestamp parse error: %s (%s)", ts_raw, e)
            continue

        doc_id = generate_snapshot_doc_id(ts, market_id)
        doc = {
            "_id": doc_id,
            "timestamp_utc": ts.isoformat(),
            "market_id": market_id,
            "question": question,
            "yes_price": round(yes_price, 6),
            "no_price": round(no_price, 6),
            "yes_cents": yes_cents,
            "no_cents": no_cents,
            "spread": spread,
            "volumeNum": volume,
            "liquidityNum": liquidity,
            "active": active,
            "closed": closed,
            "market_slug": market_slug,
        }
        docs.append(doc)

    logger.info("Parsed %d snapshots from spreadsheet", len(docs))
    return docs


def read_markets(wb) -> list[dict]:
    """Read the markets sheet into market documents."""
    ws = wb["markets"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    logger.info("markets headers: %s", headers)
    docs = []
    now = datetime.now(timezone.utc).isoformat()

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        market_id = str(int(row[0])) if isinstance(row[0], float) else str(row[0])
        market_slug = str(row[1] or "")
        question = str(row[2] or "")

        # Parse outcomes: might be JSON string like '["Yes", "No"]'
        outcomes_raw = row[3] if len(row) > 3 else "[]"
        try:
            outcomes = json.loads(str(outcomes_raw)) if outcomes_raw else []
        except (json.JSONDecodeError, TypeError):
            outcomes = []

        active = row[4] if isinstance(row[4], bool) else str(row[4]).lower() == "true" if len(row) > 4 and row[4] else True
        closed = row[5] if isinstance(row[5], bool) else str(row[5]).lower() == "true" if len(row) > 5 and row[5] else False
        volume = parse_number(row[6]) if len(row) > 6 else 0.0
        liquidity = parse_number(row[7]) if len(row) > 7 else 0.0

        # Parse source_tags: might be "politics|trump" format
        source_tags_raw = str(row[8]) if len(row) > 8 and row[8] else ""
        source_tags = [t.strip() for t in source_tags_raw.split("|") if t.strip()]

        polymarket_url = str(row[9]) if len(row) > 9 and row[9] else ""

        doc = {
            "market_id": market_id,
            "market_slug": market_slug,
            "question": question,
            "outcomes": outcomes,
            "active": active,
            "closed": closed,
            "volumeNum": volume,
            "liquidityNum": liquidity,
            "source_tags": source_tags,
            "polymarket_url": polymarket_url,
            "first_seen_utc": now,
            "last_seen_utc": now,
        }
        docs.append(doc)

    logger.info("Parsed %d markets from spreadsheet", len(docs))
    return docs


def read_tracked_markets(wb) -> list[dict]:
    """Read the tracked_markets sheet."""
    ws = wb["tracked_markets"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = rows[0]
    logger.info("tracked_markets headers: %s", headers)
    docs = []
    now = datetime.now(timezone.utc).isoformat()

    for row in rows[1:]:
        if not row or not row[0]:
            continue
        market_id = str(int(row[0])) if isinstance(row[0], float) else str(row[0])
        is_tracked = str(row[1]).upper() == "TRUE" if row[1] else True
        priority = int(row[2]) if row[2] and str(row[2]).strip() else None
        title_override = str(row[3]) if row[3] and str(row[3]).strip() else None
        notes = str(row[4]) if len(row) > 4 and row[4] and str(row[4]).strip() else None

        doc = {
            "market_id": market_id,
            "is_tracked": is_tracked,
            "priority": priority,
            "title_override": title_override,
            "notes": notes,
            "created_at_utc": now,
            "updated_at_utc": now,
        }
        docs.append(doc)

    logger.info("Parsed %d tracked markets from spreadsheet", len(docs))
    return docs


async def import_all(filepath: str):
    """Import all data from the spreadsheet into Elasticsearch."""
    logger.info("Opening %s", filepath)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    es = AsyncElasticsearch(hosts=[ES_HOST], request_timeout=60)

    try:
        # 1. Import markets (upsert by market_id)
        markets = read_markets(wb)
        if markets:
            # Check which markets already exist to preserve first_seen_utc
            for m in markets:
                try:
                    existing = await es.get(index="markets", id=m["market_id"])
                    m["first_seen_utc"] = existing["_source"].get("first_seen_utc", m["first_seen_utc"])
                except Exception:
                    pass

            actions = [{"_index": "markets", "_id": m["market_id"], "_source": m} for m in markets]
            success, errors = await async_bulk(es, actions, chunk_size=500, raise_on_error=False)
            logger.info("Markets imported: %d success, %d errors", success, len(errors) if errors else 0)

        # 2. Import snapshots (use deterministic doc_id for dedup)
        snapshots = read_snapshots_wide(wb)
        if snapshots:
            actions = []
            for s in snapshots:
                doc_id = s.pop("_id")
                actions.append({"_index": "snapshots_wide", "_id": doc_id, "_source": s})
            success, errors = await async_bulk(es, actions, chunk_size=500, raise_on_error=False)
            logger.info("Snapshots imported: %d success, %d errors", success, len(errors) if errors else 0)

        # 3. Import tracked markets (upsert by market_id)
        tracked = read_tracked_markets(wb)
        if tracked:
            actions = [{"_index": "tracked_markets", "_id": t["market_id"], "_source": t} for t in tracked]
            success, errors = await async_bulk(es, actions, chunk_size=500, raise_on_error=False)
            logger.info("Tracked markets imported: %d success, %d errors", success, len(errors) if errors else 0)

        # Force refresh so data is searchable immediately
        await es.indices.refresh(index="markets,snapshots_wide,tracked_markets")
        logger.info("Import complete!")

    finally:
        wb.close()
        await es.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <path-to-xlsx>")
        sys.exit(1)

    asyncio.run(import_all(sys.argv[1]))
